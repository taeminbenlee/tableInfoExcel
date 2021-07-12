print("hello")

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

print('테이블 명세서 코드')

#엑셀 파일 불러오기
#데이터 처리
dbInfo = pd.read_excel('saleviscaretable.xlsx')
print(dbInfo.head())
#참조테이블 찾아서 삽입
dbInfo['참조테이블'] = ''
dfFK = dbInfo[dbInfo['KEY']=='MUL']
dfPK = dbInfo[dbInfo['KEY']=='PRI']

arrFKNames = dfFK['컬럼명'].to_numpy()
#print(arrFKNames)

#PK 로 설정된 컬럼명들에서 FK로 설정된 컬럼명과 같은 컬럼들만 뽑아낸다..
#이러면 그 결과의 테이블명을 알아냄으로써 FK인 컬럼들의 참조테이블명을 알아내기 위함.
pkNames = dfPK.loc[dfPK['컬럼명'].isin(arrFKNames)]

arrParentTableNum =[]
arrParentTableNum = dfFK['컬럼명'].isin(pkNames['컬럼명']).index
#arrParentTableNum
#FK로 설정된 컬럼들의 인덱스 넘버..

#df[df['id'].isin(['b', 'e', 'k'])]

#arrParentTableNum 을 포문으로 돌려서 그 인덱스에 값을 넣어준다.

for x in arrParentTableNum:
    #fk 인덱스
    print(x)
    #컬럼명 값 얻기
    xx =[]
    xx = dfFK.loc[x, ['컬럼명']].values
    print(xx)
    #얻은 컬럼명을 PK만 있는 데이터프레임에 비교하여 인덱스 얻기
    
    xxx=dfPK.loc[dfPK['컬럼명'].isin(xx)].index
    print(xxx)
    # 얻은 인덱스로 테이블명 얻기
    #for y in xxx:
        #print(y)
    xxxx = str(dfPK.loc[xxx,['테이블명']].values)
    """
    characters = "]'["
    #문자열 제거... 
    for x in range(len(characters)):
        xxxx = xxxx.replace(characters[x],"")
        """
    print(xxxx)
        #얻은 테이블명을 참조테이블에 넣기 
      # df.loc[rowIndex, 'New Column Title'] = "some value"
      
    dbInfo.loc[x, '참조테이블'] = xxxx
    
    print(dbInfo.iloc[x]['참조테이블'])




## 반복문을 위한 테이블명을 조사 중복을 없앤다
dbInfo['테이블명'].duplicated()
## true == 중복, false == 최초의 값, 즉, false만 뽑으면 각 테이블명만 담아낼수 있다..

dbTableNames = dbInfo['테이블명'].drop_duplicates()


arrTableNames = np.array(dbTableNames.values.tolist())
print("------어레이에 담긴 테이블명들 확인-------")
print(arrTableNames)
print("----------------------------------------")
##테이블명의 인덱스 넘버
tableNameIndex = dbTableNames.index
#테이블명의 인덱스로 해당 테이블명의 논리테이블명을 가져온다
arrTableInfo =[]
for i in tableNameIndex:
    x = dbInfo.at[i, '논리테이블명']
    arrTableInfo.append(x)


# Workbook == 엑셀파일전체 AND Worksheet == 엑셀시트하나하나 
# spreadsheet 라는 변수에 .active 사용시 활성화된 시트를 불러온다. 초기엔 하나뿐..
# 시트이름 바꾸기 ex) sheet2.title = '수집 데이터'

# 원하는 셀에 값 입력하기.. 예) A1, B1, C1, D1 == '테이블 이름'

## 데이터프레임처럼 pd.read_excel 로도 엑셀파일을 불러올 수 있지만, 
## Workbook 의 메소드로도 기존 엑셀파일을 로드 할수 있다.
## 예) openpyxl.load_workbook(파일명) ==> wb=openpyxl.load_workbook('test.xlsx')
## 엑셀파일 저장은 Workbook.save(파일명) 으로 가능하다.
## 예) wb.save('test.xlsx')

## 엑셀 파일 만들기


wb = openpyxl.Workbook()
def createWB(a):
    
    print(str(a) + "번째 엑셀시트 생성")

    lenarr2 = len(arrTableNames)
    for aa in range(lenarr2):
        #aa = str(aa)
        aa = wb.create_sheet('sheet' + str(a))
        aa.title = str(a+1)
        print(str(a) + "번째 엑셀시트의 헤드 작성")
        printTableHead(aa)
        print(str(a) + "번째 엑셀시트의 값 작성")
        printTableValues(aa, a)
        break
   
    
    ##엑셀파일 저장
    print("엑셀파일 저장")
    wb.save(filename="테이블별 명세서.xlsx")
    

## 테이블명세서의 헤더 만드는 함수.

def printTableHead(aa):

    aa.merge_cells('A1:D1') 
    aa['A1']= '테이블 이름'
    aa.merge_cells('E1:J1')
    aa.merge_cells('A2:D2') 
    aa['A2']= '테이블 설명'
    aa.merge_cells('E2:J2')
    aa.merge_cells('A3:D3') 
    aa['A3']= 'PRIMARY KEY'
    aa.merge_cells('E3:J3') 
    aa.merge_cells('A4:D4') 
    aa['A4']= 'FOREIGN KEY'
    aa.merge_cells('A5:D5') 
    aa['A5']= 'INDEX'
    aa.merge_cells('A6:D6') 
    aa['A6']= 'UNIQUE INDEX'
    aa.merge_cells('E4:J4')
    aa.merge_cells('E5:J5') 
    aa.merge_cells('E6:J6') 
    aa['A7']= 'NO'
    aa.merge_cells('B7:C7')
    aa['B7']= 'PK/AI/FK'
    aa['D7']= 'NULL'
    aa.merge_cells('E7:F7')
    aa['E7']= '컬럼 이름'
    #aa['F7']= '컬럼 이름'
    aa['G7']= 'TYPE'
    aa['H7']= 'DEFAULT'
    aa['I7']= '설명'
    aa['J7']= '참조 테이블'






##테이블의 값 넣는 함수
def printTableValues(aa, a):
    print(str(a) + "값 insert 시작")

    i = arrTableNames[a]
    j = arrTableInfo[a]
    #테이블명 입력
    aa['E1'] = i
    #논리테이블명입력
    aa['E2'] = j
    ## 각 테이블명의 각각의 데이터프레임 새로 생성
    tableDetailByName = dbInfo.loc[dbInfo['테이블명'] == arrTableNames[a]]
    #PK조사
    thisPK = tableDetailByName.loc[tableDetailByName['KEY']=='PRI']
    PKString = thisPK['컬럼명'].values
    yStr=''
    if not PKString.all():
        print("PK 컬럼 없음")
    else:
        lenPK = len(PKString)
        for y in range(lenPK):
            yStr += thisPK['컬럼명'].values[y]

        #PK정보입력
        aa['E3']= yStr
    #FK조사
    thisFK = tableDetailByName.loc[tableDetailByName['KEY']=='MUL']
    FKString = thisFK['컬럼명'].values
    #FK정보입력
    zStr=''
    if not FKString.all():
        print("FK컬럼 없음")
    else:
        lenFK = len(FKString)
        for z in range(lenFK):
            zStr += thisFK['컬럼명'].values[z] + ' '
        aa['E4']= zStr

    #컬럼의 갯수조사 인덱스 설정
    nn = len(tableDetailByName.index)
    nArrNum=[]
    for nNum in range(nn):
        nArrNum += [nNum + 1]

    print(nArrNum)   

    #배경색 설정
    totalCellRows = nn+8
    my_bgColor = openpyxl.styles.colors.Color(rgb='00C0C0C0')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_bgColor)
    aa['A1'].fill = my_fill
    aa['A2'].fill = my_fill 
    aa['A3'].fill = my_fill
    aa['A4'].fill = my_fill
    aa['A5'].fill = my_fill
    aa['A6'].fill = my_fill
    aa['A7'].fill = my_fill
    aa['B7'].fill = my_fill
    aa['D7'].fill = my_fill
    aa['E7'].fill = my_fill
    aa['G7'].fill = my_fill
    aa['H7'].fill = my_fill
    aa['I7'].fill = my_fill
    aa['J7'].fill = my_fill

    #컬럼 인덱스 입력
    for i, value in enumerate(nArrNum):
        aa.cell(row=i+8, column=1, value=value)
    #컬럼명 조사
    columnNames = tableDetailByName['컬럼명']
    cn = columnNames.values
    #컬럼명 입력
    for i, value in enumerate(cn):
        aa.cell(row=i+8, column=5, value=value)
    #데이터 형 조사
    dataTypes = tableDetailByName['데이터 길이']
    dt = dataTypes.values
    #데이터 형 입력
    for i, value in enumerate(dt):
        aa.cell(row=i+8, column=7, value=value)
    #컬럼설명 조사
    columnDescs = tableDetailByName['컬럼설명']
    cd = columnDescs.values
    #컬럼설명 입력
    for i, value in enumerate(cd):
        aa.cell(row=i+8, column=9, value=value)
    #눌 허용 조사
    isNull = tableDetailByName['NULL허용']
    isn = isNull.values
    # 눌 허용값 입력
    for i, value in enumerate(isn):
        aa.cell(row=i+8, column=4, value=value)
    #디폴트값 조사
    isDefault = tableDetailByName['디폴트값']
    isd = isDefault.values
    #isdS = isd.astype(str)

    #디폴트값 입력
    for i, value in enumerate(isd):
        aa.cell(row=i+8, column=8, value=value)
    
    ## PK와 FK 컬럼에 'Y' 표기
    ## .values 또는 to._numpy()  numpy배열로 변환함!
    ## 인덱싱으로 특정값을 변경한다. (Indexing: a[a < 0] = 0) 
    isThisPK = tableDetailByName['KEY'].to_numpy()
    isThisPK[isThisPK=='MUL']='FK'
    for i, value in enumerate(isThisPK):
        aa.cell(row=i+8, column=2, value=value)
    
    #참조테이블 정보 입력

    isThisParentTable = tableDetailByName['참조테이블']
    isTPT = isThisParentTable.values
    for i, value in enumerate(isTPT):
        aa.cell(row=i+8, column=10, value=value)

    #보더 설정
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bb = len(aa['J'])
    bbb = 'J'+str(bb)
    for row in aa['A1':bbb]:
        for cell in row:
            cell.border = border









#테이블명의 length
lenarr = len(arrTableNames)

#for loop 을 이용하여 함수호출!

print("시트생성 작업 시작")
for a in range(lenarr):
    
    createWB(a)
print("시트생성 작업 끝")


print("모든 sheet combine 작업 시작")

# empty dictionary 생성
df = []
# 여러 시트를 가지고 있는 엑셀파일
#wb = openpyxl.load_workbook('테이블명세서.xlsx')  같은 방법.
f = "D:\dev-envs\pyexcel/테이블별 명세서.xlsx" 

# 합치고자 하는 시트의 갯수..
#numberOfSheets = 477
xl = pd.ExcelFile('D:\dev-envs\pyexcel/테이블별 명세서.xlsx')
numberOfSheets = len(xl.sheet_names)

print("합칠 시트의수: "+str(numberOfSheets))

for i in range(1,numberOfSheets):    
     data = pd.read_excel(f, sheet_name = str(i), header=None) 
     df.append(data)
#remember python is very strict on how you arrange stuff so be aware of this

#새로 저장할 엑셀파일의 저장 경로와 이름 지정
print('새로 저장할 엑셀파일의 저장 경로와 이름 지정')
final = "D:\dev-envs\pyexcel//테이블명세서.xlsx"
print("D:\dev-envs\pyexcel//테이블명세서.xlsx")

#모든 시트를 담은 데이터프레임을 합친다
df = pd.concat(df)

#데이터프레임의 인덱스와 헤더를 제거한후 엑셀로 저장한다.
df.to_excel(final, sheet_name='테이블명세서', index=False, header=False)

#보더 설정
from openpyxl import load_workbook
wb1 = load_workbook(filename = '테이블명세서.xlsx')
ws1 = wb1.active

thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
bb = len(ws1['J'])
bbb = 'J'+str(bb)
for row in ws1['A1':bbb]:
    for cell in row:
        cell.border = border


print("테이블명세서 작업 완료")



